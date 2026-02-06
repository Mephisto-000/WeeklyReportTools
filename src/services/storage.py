from pathlib import Path
from datetime import date
from typing import List, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from src.models.record import WorkRecord
from src.utils.logger import setup_logger

logger = setup_logger("Storage")

class ExcelStorage:
    """Handles persistence of work records to an Excel file."""

    FILE_NAME = "work_reports.xlsx"
    SHEET_NAME = "WorkRecords"
    HEADERS = ["Date", "Project", "Summary", "Details"]

    def __init__(self, file_path: str = FILE_NAME):
        self.file_path = Path(file_path)
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Creates the Excel file with headers if it doesn't exist."""
        if not self.file_path.exists():
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.SHEET_NAME
            sheet.append(self.HEADERS)
            workbook.save(self.file_path)
            logger.info(f"Created new storage file: {self.file_path}")

    def _get_workbook_sheet(self):
        """Opens the workbook and returns the active sheet."""
        workbook = openpyxl.load_workbook(self.file_path)
        if self.SHEET_NAME in workbook.sheetnames:
            sheet = workbook[self.SHEET_NAME]
        else:
            sheet = workbook.create_sheet(self.SHEET_NAME)
            sheet.append(self.HEADERS)
        return workbook, sheet

    def save_record(self, record: WorkRecord):
        """Appends a new record to the Excel file."""
        workbook, sheet = self._get_workbook_sheet()
        
        # Check if updating an existing record logic is needed?
        # Requirement says "Edit existing record"
        # To support edit, we need to identify records. 
        # Since we don't have IDs, we might delete-then-insert or find-and-update.
        # Given the requirements, I will implement a safe append for now 
        # and handle "Edit" as "Delete + Add" or row replacement in the GUI logic 
        # or add a specific update method here.
        # Let's add an internal ID or just rely on row index for editing.
        # Ideally, we should use row index for deletion/editing.
        
        row_data = [
            record.date.isoformat(),
            record.project_name,
            record.summary,
            record.details
        ]
        sheet.append(row_data)
        workbook.save(self.file_path)
        logger.info(f"Saved record: {record.summary}")

    def get_records_by_date(self, target_date: date) -> List[tuple[int, WorkRecord]]:
        """
        Retrieves records for a specific date.
        Returns a list of tuples: (row_index, WorkRecord).
        Row index is 1-based (openpyxl style).
        """
        workbook, sheet = self._get_workbook_sheet()
        records = []
        
        # Iterate rows, skipping header
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]: continue
            
            try:
                record_date_str = str(row[0]).split("T")[0] # handle datetime objects or strings
                if record_date_str == target_date.isoformat():
                    record = WorkRecord(
                        date=date.fromisoformat(record_date_str),
                        project_name=str(row[1]) if row[1] else "",
                        summary=str(row[2]) if row[2] else "",
                        details=str(row[3]) if row[3] else ""
                    )
                    records.append((i, record))
            except Exception as e:
                logger.error(f"Error parsing row {i}: {e}")
                continue
                
        return records

    def delete_record_by_row(self, row_index: int):
        """Deletes a record specified by its Excel row index."""
        workbook, sheet = self._get_workbook_sheet()
        
        # Openpyxl delete_rows is 1-based? Yes.
        sheet.delete_rows(row_index)
        workbook.save(self.file_path)
        logger.info(f"Deleted record at row {row_index}")

    def update_record_by_row(self, row_index: int, record: WorkRecord):
        """Updates a record at a specific row."""
        workbook, sheet = self._get_workbook_sheet()
        
        sheet.cell(row=row_index, column=1, value=record.date.isoformat())
        sheet.cell(row=row_index, column=2, value=record.project_name)
        sheet.cell(row=row_index, column=3, value=record.summary)
        sheet.cell(row=row_index, column=4, value=record.details)
        
        workbook.save(self.file_path)
        logger.info(f"Updated record at row {row_index}")
